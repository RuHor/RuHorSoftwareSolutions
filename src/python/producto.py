# Copyright (c) 2024, [RuHor]
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import sys
import os
import openpyxl
from PySide6.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QLabel, QPushButton, QLineEdit, QFileDialog, QMessageBox, QTextEdit, QWidget
from PySide6.QtCore import Qt
from PySide6.QtGui import QIcon

class ExcelSearchApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Buscar productos')

        self.layout = QVBoxLayout()

        self.label = QLabel('Termino a buscar:')
        self.layout.addWidget(self.label)

        self.edit = QLineEdit()
        self.layout.addWidget(self.edit)

        self.file_label = QLabel('Archivo seleccionado: ')
        self.layout.addWidget(self.file_label)

        self.select_button = QPushButton('Seleccionar archivo Excel')
        self.select_button.clicked.connect(self.select_excel_file)
        self.layout.addWidget(self.select_button)

        self.open_button = QPushButton('Abrir Archivo')
        self.open_button.clicked.connect(self.open_selected_file)
        self.layout.addWidget(self.open_button)

        self.search_button = QPushButton('Buscar')
        self.search_button.clicked.connect(self.search_text)
        self.layout.addWidget(self.search_button)

        self.result_text_edit = QTextEdit()
        self.result_text_edit.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.result_text_edit.setLineWrapMode(QTextEdit.NoWrap)
        self.layout.addWidget(self.result_text_edit)

        central_widget = QWidget()
        central_widget.setLayout(self.layout)
        self.setCentralWidget(central_widget)

        self.line_count_label = QLabel('Coincidencia(s) encontrada(s): 0')
        self.layout.addWidget(self.line_count_label)

        self.selected_file = None

    def select_excel_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly

        file_name, _ = QFileDialog.getOpenFileName(self, "Seleccionar archivo Excel", "", "Archivos Excel (*.xlsx);;Todos los archivos (*)", options=options)

        if file_name:
            self.selected_file = file_name
            self.file_label.setText('Archivo seleccionado: ' + file_name)

    def open_selected_file(self):
        if self.selected_file:
            try:
                os.startfile(self.selected_file)
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Error al abrir el archivo: {e}')
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se ha seleccionado ningún archivo.')

    def search_text(self):
        if self.selected_file is None:
            QMessageBox.warning(self, 'Error', 'Por favor, selecciona un archivo Excel primero.')
            return

        search_text = self.edit.text()

        if not search_text:
            QMessageBox.warning(self, 'Error', 'Por favor, ingresa un texto para buscar.')
            return

        try:
            workbook = openpyxl.load_workbook(self.selected_file)
            found_results = []

            for sheet in workbook:
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        if search_text.lower() in str(cell).lower():
                            cell_value = str(cell)
                            cell_address = sheet.cell(row=row_idx, column=col_idx).coordinate
                            next_col_value = str(sheet.cell(row=row_idx, column=col_idx + 1).value)

                            next_col_value_no_dot = next_col_value.replace(".", "")
                            if next_col_value_no_dot.isnumeric():
                                next_col_value = float(sheet.cell(row=row_idx, column=col_idx + 1).value)
                                next_col_value1 = next_col_value * 0.12
                                next_col_value1 = round(next_col_value + next_col_value1, 2)
                                found_results.append(f'{cell_address} : {cell_value}  -  ($ {next_col_value}  sin iva) -  ($ {next_col_value1} con iva)')
                            else:
                                next_col_value = str(sheet.cell(row=row_idx, column=col_idx + 1).value)
                                found_results.append(f'{cell_address} : {cell_value}')

            if found_results:
                self.result_text_edit.setPlainText('\n'.join(found_results))
                line_count = self.result_text_edit.toPlainText().count('\n')
                self.line_count_label.setText(f'Coincidencia(s) encontrada(s): {line_count}')

                if found_results:
                    self.result_text_edit.setPlainText('\n'.join(found_results))
                else:
                    self.result_text_edit.setPlainText('No se encontraron coincidencias.')
                    self.line_count_label.setText(f'Coincidencia(s) encontrada(s): 0')
            else:
                self.result_text_edit.setPlainText('No se encontraron coincidencias.')
                self.line_count_label.setText(f'Coincidencia(s) encontrada(s): 0')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Error al abrir el archivo Excel: {e}')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ExcelSearchApp()
    icon = QIcon("icon.ico") 
    window.setWindowIcon(icon)
    window.show()
    sys.exit(app.exec())
